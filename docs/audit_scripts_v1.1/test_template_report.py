"""端到端测试：模板格式报告生成"""
import sys, warnings
warnings.filterwarnings('ignore')
sys.path.insert(0, '/workspace')

import pandas as pd
from datetime import date
from src.report.record_builder import RecordBuilder
from src.report.excel_report import ExcelReporter

# ===== 模拟 Tier1 雷达池数据 =====
tier1 = pd.DataFrame([
    {'symbol': '000651', 'name': '格力电器', 'industry': '家用电器',
     'price': 40.5, 'pe_dynamic': 8.2, 'pb': 1.6, 'market_cap': 2280e8,
     'change_60d': -22.5, 'pe_percentile': 12.0, 'dividend_yield': 5.8},
    {'symbol': '600519', 'name': '贵州茅台', 'industry': '食品饮料',
     'price': 1450.0, 'pe_dynamic': 20.5, 'pb': 8.9, 'market_cap': 18200e8,
     'change_60d': -15.2, 'pe_percentile': 25.0, 'dividend_yield': 2.1},
    {'symbol': '601088', 'name': '中国神华', 'industry': '煤炭',
     'price': 38.6, 'pe_dynamic': 11.2, 'pb': 1.8, 'market_cap': 7670e8,
     'change_60d': -18.3, 'pe_percentile': 18.0, 'dividend_yield': 6.5},
    {'symbol': '600036', 'name': '招商银行', 'industry': '银行',
     'price': 33.8, 'pe_dynamic': 6.1, 'pb': 0.9, 'market_cap': 8530e8,
     'change_60d': -12.8, 'pe_percentile': 22.0, 'dividend_yield': 5.2},
])

# ===== 模拟 Tier2 深度观察池数据 =====
tier2 = pd.DataFrame([
    {'symbol': '000651', 'name': '格力电器', 'industry': '家用电器',
     'price': 40.5, 'pe_dynamic': 8.2, 'pb': 1.6, 'market_cap': 2280e8,
     'change_60d': -22.5, 'pe_percentile': 12.0, 'dividend_yield': 5.8,
     'total_score': 7.2, 'odds_ratio': 2.35, 'confidence': 0.72,
     'implied_roe': 19.5,
     'dimension_scores': {'business_quality': 8, 'competitive_advantage': 7,
                          'demand_certainty': 6, 'management': 6,
                          'financial_quality': 8, 'valuation_margin': 9,
                          'odds': 8, 'predictability': 7,
                          'market_pessimism': 8, 'reversal_verifiability': 6},
     'overall_risk_level': '低'},
    {'symbol': '601088', 'name': '中国神华', 'industry': '煤炭',
     'price': 38.6, 'pe_dynamic': 11.2, 'pb': 1.8, 'market_cap': 7670e8,
     'change_60d': -18.3, 'pe_percentile': 18.0, 'dividend_yield': 6.5,
     'total_score': 6.8, 'odds_ratio': 1.95, 'confidence': 0.68,
     'implied_roe': 16.1,
     'dimension_scores': {'business_quality': 7, 'competitive_advantage': 7,
                          'demand_certainty': 5, 'management': 7,
                          'financial_quality': 8, 'valuation_margin': 8,
                          'odds': 6, 'predictability': 6,
                          'market_pessimism': 7, 'reversal_verifiability': 6},
     'overall_risk_level': '低'},
])

# ===== 模拟 Tier3 核心黄金坑数据 =====
tier3 = pd.DataFrame([
    {'symbol': '000651', 'name': '格力电器', 'industry': '家用电器',
     'price': 40.5, 'pe_dynamic': 8.2, 'pb': 1.6, 'market_cap': 2280e8,
     'change_60d': -22.5, 'pe_percentile': 12.0, 'dividend_yield': 5.8,
     'total_score': 7.2, 'odds_ratio': 2.35, 'confidence': 0.72,
     'implied_roe': 19.5,
     'dimension_scores': {'business_quality': 8, 'competitive_advantage': 7,
                          'demand_certainty': 6, 'management': 6,
                          'financial_quality': 8, 'valuation_margin': 9,
                          'odds': 8, 'predictability': 7,
                          'market_pessimism': 8, 'reversal_verifiability': 6},
     'overall_risk_level': '低'},
])

# ===== 生成报告 =====
reporter = ExcelReporter()
builder = RecordBuilder(scan_date=date.today())

# 组装完整记录
t2_records = []
for _, row in tier2.iterrows():
    rec = builder.build_full_record(
        row,
        valuation={},
        expectation={},
        scores=row['dimension_scores'],
        risk={'overall_risk_level': row['overall_risk_level']},
        rating='B',
        odds_info={'odds_ratio': row['odds_ratio']},
        financial={'roe': row['implied_roe'], 'dividend_yield': row['dividend_yield']},
    )
    t2_records.append(rec)

t3_records = []
for _, row in tier3.iterrows():
    rec = builder.build_full_record(
        row,
        valuation={},
        expectation={},
        scores=row['dimension_scores'],
        risk={'overall_risk_level': row['overall_risk_level']},
        rating='A',
        odds_info={'odds_ratio': row['odds_ratio']},
        financial={'roe': row['implied_roe'], 'dividend_yield': row['dividend_yield']},
    )
    t3_records.append(rec)

# 证伪日志
fal_logs = []
for _, row in tier3.iterrows():
    conds = [
        {'condition': 'ROE(TTM)跌破阈值', 'threshold': '8%',
         'current': f"{row['implied_roe']}%", 'triggered': False},
        {'condition': '经营现金流/净利润<50%', 'threshold': '50%',
         'current': '待人工核对', 'triggered': False},
    ]
    fal_logs.extend(builder.build_falsification_log(row, conds))

# 价值陷阱（示例）
traps = [builder.build_trap_record(
    pd.Series({'symbol': '000002', 'name': '万科A', 'industry': '房地产'}),
    reason='地产需求中枢下移，净利润连续亏损，股息可持续性存疑',
    evidence='2025Q1归母净利-62.5亿；行业销售面积连续4年下滑')]

# 生成
path = reporter.generate_full_report(
    scan_date=date.today(),
    tier1=tier1,
    tier2=t2_records,
    tier3=t3_records,
    traps=traps,
    falsification_logs=fal_logs,
)
print(f'生成: {path}')

# ===== 验证 =====
import openpyxl
wb = openpyxl.load_workbook(path)
print(f'Sheets: {wb.sheetnames}')
for name in wb.sheetnames:
    ws = wb[name]
    print(f'  [{name}] {ws.max_row}行 x {ws.max_column}列')

# 抽查核心黄金坑的内容
ws = wb['核心黄金坑']
print('\n=== 核心黄金坑 第3行（格力电器）===')
headers = [ws.cell(row=2, column=j).value for j in range(1, ws.max_column+1)]
values = [ws.cell(row=3, column=j).value for j in range(1, ws.max_column+1)]
for h, v in zip(headers, values):
    if v is not None and v != '':
        print(f'  {h}: {v}')

# 抽查证伪日志
ws = wb['证伪日志']
print('\n=== 证伪日志 ===')
for i in range(1, ws.max_row+1):
    row_vals = [ws.cell(row=i, column=j).value for j in range(1, 8)]
    print(f'  {row_vals}')

# 验证下拉验证是否保留
ws = wb['深度观察池']
print(f'\n深度观察池下拉验证规则数: {len(ws.data_validations.dataValidation)}')
