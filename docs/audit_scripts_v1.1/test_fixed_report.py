"""修正后的端到端测试：使用真实财务数据验证"""
import sys, warnings
warnings.filterwarnings('ignore')
sys.path.insert(0, '/workspace')

import pandas as pd
from datetime import date
from src.report.record_builder import RecordBuilder
from src.report.excel_report import ExcelReporter

# ===== 模拟真实扫描数据（格力，含真实财务）=====
tier1 = pd.DataFrame([
    {'symbol': '000651', 'name': '格力电器', 'industry': '家用电器',
     'price': 40.5, 'pe_dynamic': 8.2, 'pb': 1.6, 'market_cap': 2280e8,
     'change_60d': -22.5, 'pe_percentile': 12.0, 'dividend_yield': 5.8},
])

tier3 = pd.DataFrame([
    {'symbol': '000651', 'name': '格力电器', 'industry': '家用电器',
     'price': 40.5, 'pe_dynamic': 8.2, 'pb': 1.6, 'market_cap': 2280e8,
     'change_60d': -22.5, 'pe_percentile': 12.0, 'dividend_yield': 5.8,
     'total_score': 7.2, 'odds_ratio': None, 'confidence': 0.72,
     'implied_roe': 19.5,
     'dimension_scores': {'business_quality': 8, 'competitive_advantage': 7,
                          'demand_certainty': 6, 'management': 6,
                          'financial_quality': 8, 'valuation_margin': 9,
                          'odds': 8, 'predictability': 7,
                          'market_pessimism': 8, 'reversal_verifiability': 6},
     'overall_risk_level': '低'},
])

builder = RecordBuilder(scan_date=date.today())

# 格力的真实正常化利润（2025年报归母净利约290亿，保守取260亿）
gree_normalized_profit = 260  # 亿元

t3_records = []
for _, row in tier3.iterrows():
    rec = builder.build_full_record(
        row,
        valuation={},
        expectation={},
        scores=row['dimension_scores'],
        risk={'overall_risk_level': row['overall_risk_level']},
        rating='A',
        odds_info={},
        financial={
            'roe': 19.5,
            'dividend_yield': 5.8,
            'normalized_profit': gree_normalized_profit,  # 真实正常化利润
        },
    )
    t3_records.append(rec)

reporter = ExcelReporter()
path = reporter.generate_full_report(
    scan_date=date.today(),
    tier1=tier1,
    tier2=[],
    tier3=t3_records,
    traps=[],
    falsification_logs=[],
)
print(f'生成: {path}')

# ===== 验证核心字段 =====
import openpyxl
wb = openpyxl.load_workbook(path)
ws = wb['核心黄金坑']
headers = [ws.cell(row=2, column=j).value for j in range(1, ws.max_column+1)]
values = [ws.cell(row=3, column=j).value for j in range(1, ws.max_column+1)]

print('\n=== 关键字段验证（格力电器）===')
key_fields = ['总市值(亿元)', 'PE(TTM)', '市场隐含利润(亿元)', '隐含利润反推口径',
              '我的基准假设-正常化利润(亿元)', '预期差(基准÷隐含)',
              '悲观估值(亿元)', '合理估值(亿元)', '乐观估值(亿元)',
              '潜在下跌空间(%)', '合理上涨空间(%)', '赔率(合理估值÷市值)',
              '概率(低/中/高)', '赔率等级(低/中/高/极高)', '仓位适配度']
for h, v in zip(headers, values):
    if h in key_fields:
        print(f'  {h}: {v}')

# 手工验算
print('\n=== 手工验算 ===')
mc = 2280
pe = 8.2
implied = mc / pe
print(f'市场隐含利润 = {mc}÷{pe} = {implied:.1f}亿')
gap = gree_normalized_profit / implied
print(f'预期差 = {gree_normalized_profit}÷{implied:.1f} = {gap:.2f}')
base_val = gree_normalized_profit * 15
print(f'合理估值 = {gree_normalized_profit}×15 = {base_val}亿')
odds = base_val / mc
print(f'赔率 = {base_val}÷{mc} = {odds:.2f}')
upside = (base_val / mc - 1) * 100
print(f'合理上涨空间 = {upside:.1f}%')
