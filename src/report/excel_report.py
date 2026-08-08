"""
Excel 报告生成器（模板填充版）

以用户设计的《黄金坑数据库》模板为基底：
- 保留「使用说明」「字段字典」两个静态工作表
- 保留所有分组表头、合并单元格、下拉验证、格式
- 自动填充：雷达池 / 深度观察池 / 核心黄金坑 / 价值陷阱 / 证伪日志

字段策略：
- 可量化字段由系统自动计算（估值、预期差、赔率、评分、证伪读数）
- 人工判断字段给出系统初稿并标注「待人工复核」
- 每条记录自动标注数据日期、口径、来源
"""

import logging
import shutil
from datetime import date
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .record_builder import (
    FULL_COLUMNS,
    RADAR_COLUMNS,
    TRAP_COLUMNS,
    FALSIFICATION_COLUMNS,
    RecordBuilder,
)

logger = logging.getLogger(__name__)

# 模板默认路径
DEFAULT_TEMPLATE = Path(__file__).parent.parent.parent / 'assets' / '黄金坑数据库_模板.xlsx'


class ExcelReporter:
    """黄金坑数据库 Excel 报告生成器（模板填充版）

    输出工作表与用户模板完全一致：
    使用说明 / 雷达池 / 深度观察池 / 核心黄金坑 / 价值陷阱 / 证伪日志 / 字段字典
    """

    def __init__(self, db=None, output_dir: Optional[Path] = None,
                 template_path: Optional[Path] = None):
        """初始化

        Args:
            db: 数据库连接（可选）
            output_dir: 报告输出目录
            template_path: 模板文件路径，默认 assets/黄金坑数据库_模板.xlsx
        """
        self.db = db
        self.output_dir = output_dir or Path('/workspace/output/reports')
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.template_path = Path(template_path) if template_path else DEFAULT_TEMPLATE

    def generate_full_report(
        self,
        scan_date: Optional[date] = None,
        tier1: Optional[pd.DataFrame] = None,
        tier2: Optional[pd.DataFrame] = None,
        tier3: Optional[pd.DataFrame] = None,
        traps: Optional[List[Dict]] = None,
        falsification_logs: Optional[List[Dict]] = None,
        fetcher=None,
    ) -> Path:
        """生成完整黄金坑数据库 Excel

        Args:
            scan_date: 扫描日期
            tier1: 雷达池数据（DataFrame，原始行情行）
            tier2: 深度观察池数据（DataFrame 或 List[Dict]，56列记录）
            tier3: 核心黄金坑数据（DataFrame 或 List[Dict]，56列记录）
            traps: 价值陷阱记录列表（10列字典）
            falsification_logs: 证伪日志记录列表（10列字典）
            fetcher: DataFetcher（用于 RecordBuilder 补充数据）

        Returns:
            生成的 Excel 文件路径
        """
        scan_date = scan_date or date.today()
        filename = f"黄金坑数据库_{scan_date.strftime('%Y%m%d')}.xlsx"
        filepath = self.output_dir / filename

        builder = RecordBuilder(fetcher=fetcher, scan_date=scan_date)

        # 1. 复制模板（保留所有格式、说明、下拉验证）
        if self.template_path.exists():
            shutil.copy(self.template_path, filepath)
            wb = load_workbook(filepath)
        else:
            logger.warning(f"模板不存在: {self.template_path}，将创建空白结构")
            wb = self._create_fallback_workbook()

        # 2. 填充雷达池（17列轻量）
        self._fill_radar_sheet(wb, tier1, builder)

        # 3. 填充深度观察池（56列完整）
        self._fill_full_sheet(wb, '深度观察池', tier2, builder)

        # 4. 填充核心黄金坑（56列完整）
        self._fill_full_sheet(wb, '核心黄金坑', tier3, builder)

        # 5. 填充价值陷阱（10列）
        self._fill_simple_sheet(wb, '价值陷阱', TRAP_COLUMNS, traps or [])

        # 6. 填充证伪日志（10列）
        self._fill_simple_sheet(wb, '证伪日志', FALSIFICATION_COLUMNS,
                                falsification_logs or [])

        wb.save(filepath)
        logger.info(f"黄金坑数据库已生成: {filepath}")
        return filepath

    # =========================================================
    # 雷达池（数据从第2行开始，无分组表头）
    # =========================================================
    def _fill_radar_sheet(self, wb, tier1: Optional[pd.DataFrame],
                          builder: RecordBuilder):
        ws = wb['雷达池']
        # 清除模板中的空占位行（保留表头第1行）
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

        if tier1 is None or tier1.empty:
            return

        for i, (_, row) in enumerate(tier1.iterrows()):
            try:
                record = builder.build_radar_record(row)
            except Exception as e:
                logger.debug(f"雷达池记录组装失败 {row.get('symbol')}: {e}")
                continue
            for j, col in enumerate(RADAR_COLUMNS, 1):
                ws.cell(row=i + 2, column=j, value=record.get(col))

        logger.info(f"雷达池填充 {min(len(tier1), ws.max_row - 1)} 条")

    # =========================================================
    # 深度观察池/核心黄金坑（56列，数据从第3行开始）
    # =========================================================
    def _fill_full_sheet(self, wb, sheet_name: str,
                         data, builder: RecordBuilder):
        ws = wb[sheet_name]
        # 清除模板中的空占位行（保留第1行分组表头+第2行列名）
        if ws.max_row > 2:
            ws.delete_rows(3, ws.max_row - 2)

        if data is None:
            return

        # 支持 DataFrame（原始行）或 List[Dict]（已组装好的56列记录）
        records: List[Dict] = []
        if isinstance(data, pd.DataFrame):
            if data.empty:
                return
            for _, row in data.iterrows():
                try:
                    # 若行内已含组装好的字段（full_record 键），直接用
                    if 'full_record' in row and isinstance(row['full_record'], dict):
                        records.append(row['full_record'])
                    else:
                        records.append(builder.build_full_record(row))
                except Exception as e:
                    logger.debug(f"{sheet_name}记录组装失败 {row.get('symbol')}: {e}")
        elif isinstance(data, list):
            records = [r for r in data if isinstance(r, dict)]

        for i, record in enumerate(records):
            for j, col in enumerate(FULL_COLUMNS, 1):
                ws.cell(row=i + 3, column=j, value=record.get(col))

        logger.info(f"{sheet_name}填充 {len(records)} 条")

    # =========================================================
    # 简单表（价值陷阱/证伪日志，数据从第2行开始）
    # =========================================================
    def _fill_simple_sheet(self, wb, sheet_name: str,
                           columns: List[str], records: List[Dict]):
        ws = wb[sheet_name]
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

        for i, record in enumerate(records):
            for j, col in enumerate(columns, 1):
                ws.cell(row=i + 2, column=j, value=record.get(col))

        if records:
            logger.info(f"{sheet_name}填充 {len(records)} 条")

    # =========================================================
    # 兜底：模板缺失时创建简化结构
    # =========================================================
    def _create_fallback_workbook(self):
        """模板文件缺失时创建简化版工作簿"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79",
                                  fill_type="solid")
        header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")

        def setup(ws, cols):
            for j, c in enumerate(cols, 1):
                cell = ws.cell(row=1, column=j, value=c)
                cell.font = header_font
                cell.fill = header_fill
                ws.column_dimensions[get_column_letter(j)].width = 14

        wb.active.title = '雷达池'
        setup(wb['雷达池'], RADAR_COLUMNS)
        ws2 = wb.create_sheet('深度观察池')
        setup(ws2, FULL_COLUMNS)
        ws3 = wb.create_sheet('核心黄金坑')
        setup(ws3, FULL_COLUMNS)
        ws4 = wb.create_sheet('价值陷阱')
        setup(ws4, TRAP_COLUMNS)
        ws5 = wb.create_sheet('证伪日志')
        setup(ws5, FALSIFICATION_COLUMNS)
        return wb
